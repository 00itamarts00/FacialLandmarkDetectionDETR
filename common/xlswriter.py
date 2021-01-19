# conda install -c anaconda openpyxl
from os import path

import openpyxl
from openpyxl import load_workbook


class CXlsWriter:
    def __init__(self, dbxlsfile, to_append=False):
        self.dbxlsfile = dbxlsfile

        if not path.exists(self.dbxlsfile) or not to_append:
            self.wb = openpyxl.Workbook()
            self.wb.save(self.dbxlsfile)
        else:
            self.wb = load_workbook(filename=dbxlsfile)

        self.ws = self.wb.worksheets[0]
        self.curr_row = 0

    def add_row(self, data):
        self.curr_row = self.ws.max_row + 1

        for item in data:
            self.insert_value(data[item], item)

    def search_value_in_row_index(self, field, row=1):
        for cell in self.ws[row]:
            if cell.value == field:
                return cell.column, row

        return -1, -1

    def insert_value(self, value, field):
        [column_count, row_title] = self.search_value_in_row_index(field)
        if column_count < 0:
            column_count = self.ws.max_column
            column_count = column_count + 1
            self.ws.cell(row=1, column=column_count).value = field

        self.ws.cell(row=self.curr_row, column=column_count).value = value

    def dump(self):
        self.wb.save(self.dbxlsfile)
